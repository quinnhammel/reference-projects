# this is the file to use to actually call methods for scraping.
import asyncio
from pyppeteer import launch
from pyppeteer.errors import TimeoutError
from pyppeteer.errors import NetworkError
from bs4 import BeautifulSoup
from shared_parse import count_emails
from sparse_parse import scrape_all_info as sparse_scrape
from dense_parse import scrape_all_info as dense_scrape
# imports for spreadsheet stuff...
from collections import namedtuple
import os.path
from os import path
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from datetime import datetime
import sys

# DESCRIPTION:
  # this file contains functions to get the full html for a site (by scrolling to the bottom), and to read and write to google sheets for input and output.


# CONSTANTS: tweak these to change behavior of program. Warning: choosing invalid values will cause undefined behavior. As a whole, these values must be carefully chosen.

MIN_TIMEOUT_PERIOD = 5_000
MAX_TIMEOUT_PERIOD = 15_000
TIMEOUT_ATTEMPTS = 10 # the program will try to scroll to the bottom with a timeout of MIN_TIMEOUT_PERIOD. If it encounters an UNEXPECTED timeout error, it will increase the timeout and try again.
# it will do this TIMEOUT_ATTEMPTS times, until the timeout reaches MAX_TIMEOUT_PERIOD.
# NOTE: if the timeout becomes too large, causing a network error (pyppeteer will close the connection), the function will give up before reaching MAX_TIMEOUT_PERIOD

DENSE_EMAILS_THRESHOLD = 10 # number of emails that must be found in a page's soup for it to be considered 'dense'.


# excel constants:
INPUT_START_ROW = 2 # the row that the input will start on... This allows for one row, which can hold titles for columns, etc.
INPUT_UNIVERSITY_COL = 1  # the col that university names will be under.
INPUT_DEPARTMENT_COL = 2 # similar, for departments.
INPUT_LINK_COL = 3 # similar, for the links...

DEFAULT_OUTPUT_NAME = "CollegeContactData"

YELLOW_HEX = "FFFC52"
GREEN_HEX = "1ADA1A"
RED_HEX = "FF0000"
LIGHT_GRAY_HEX = "B7B7B7"

DeptLink = namedtuple("DeptLink", ["department", "link"]) # simple namedtuple for the input function.

# the following method will scrape one link. We will call either parse_scrape or dense_scrape.
# before that, however, we need to get the WHOLE site (scrolling). This is probably the sketchiest and most fragile part of the code.
# beware...
async def get_full_html(link: str):
  output = None
  # procedure: we get the html to the link and then keep trying to scroll down.
  # we scroll down and wait for the page to load. When it does not load (it will timeout), then we have reached the end of the page.
  # we then scroll to the bottom of the page and wait for the page to reload (it might not)
  # we repeatedly do this until the number of emails on the page does not change (or the request times out.)

  # get the website using pyppeteer
  # we attempt this multiple times, using varying timeouts.
  for attempt_num in range(TIMEOUT_ATTEMPTS):
    print(f"Attempting to scroll to bottom of page {link} ({attempt_num+1}/{TIMEOUT_ATTEMPTS})...")
    browser = await launch()
    page = await browser.newPage()
    try:
      timeout_period = MIN_TIMEOUT_PERIOD + (attempt_num / (TIMEOUT_ATTEMPTS - 1 )) * (MAX_TIMEOUT_PERIOD - MIN_TIMEOUT_PERIOD) # this will start as MIN, go to MAX one step at a time.
    except:
      timeout_period = MIN_TIMEOUT_PERIOD # in case we divide by 0.

    if timeout_period <= 0:
      continue # invalid timeout_period!

    page.setDefaultNavigationTimeout(timeout_period)

    # we now put the rest in a try block. If we except a NetworkError, we know that the timeout is too long, and we must stop.
    try:
      await page.goto(link)
      # we go through the elements on the page and try to scroll. We will do this, until the page cannot scroll any longer (a timeout). So, we have to use this--admittedly degenerate--while True loop.
      while True:
        # we need to grab all the elements on the page so we can try to scroll somewhere near the bottom.
        # we need to grab all the elements because you cannot always scroll to an element.
        
        all_elements = await page.querySelectorAll("*")

        # we now go through backwards, trying to hover on one (this will automatically scroll to the bottom, where that element is in view.)
        for element in reversed(all_elements):
          try:
            await (element.hover()) # this will try to hover the mouse over the element (cannot hover over some elements in websites)
            break # if we suceed, the page will'hover' the mouse. This will scroll that portion of the page into view, if not. This may or may not trigger a navigation. So, we break out to check for the navigation.
          except:
            pass # try to hover over the next element
          
        try:
          await page.waitForNavigation() # if the page is going to change, this will wait for that. IF it does reload a navigation, we need to start the scrolling process again (can scroll multiple times.)
        except TimeoutError:
          break # if we time out, it just means we are at a point where the page does not change. This means that we are done with the while True loop.

      # when we get here, the page has completely loaded.
      output = await page.content()
      await browser.close() # need to close the browser
      return output
    except TimeoutError:
      # if we get an unexpected timeout error, we should continue on to the next attempt, since the timeout length will increase.
      continue
    except NetworkError:
      print("Could not fetch html (NetworkError)!!!")
      break # when we hit a network error, we need to break out of the whole attempt loop. This is because the timout has gotten too long, and we will not be able to retrieve this site.
    except Exception:
      print("Could not fetch html (generic Exception)!!!")
        
  return output # in case something goes wrong...

 
# the following function will scrape a SINGLE link, and return the name to emails dictionary for that link. This will delegate the work to either the dense or sparse scraper, depending on if the number of emails is above or below DENSE_EMAILS_THRESHOLD (defined above).
# this function must be asynchronous because it will use get_full_html.
async def scrape_individual(link: str):
  output = {} # names and emails dictionary
  try:
    link = str(link)
  except:
    return output # link was not convertable to a str.

  # we want to try to get the html. If this fails, we want to return output ({}), but we also want to print the exception for logging purposes.
  try:
    html = await get_full_html(link)
  except Exception as e:
    print(f"Problem with getting full html from link, raised: {str(e)}")
    return output
  
  # simple case: if html is None, we could not fetch the link.
  if html is None:
    return output

  # we now proceed similarly for translating the html to soup.
  try:
    soup = BeautifulSoup(html, "html.parser")
  except Exception as e:
    print(f"Problem with creating soup from html, raised: {str(e)}")

  # we want to consider whether to delegate this to the sparse scraper or the dense scraper.
  if count_emails(soup) >= DENSE_EMAILS_THRESHOLD:
    print("Starting individual dense parsing...")
    output = dense_scrape(soup) # want to use the dense scraping.
  else:
    print("Starting individual sparse parsing...")
    output = sparse_scrape(link, soup)
  
  return output

async def test(link: str):
  print(link)
  try:
    names_and_emails = await scrape_individual(link)
  except Exception as e:
    print(e)
    print("uhoh!")
    return 
  print("\n\n")
  for name, email in names_and_emails.items():
    print(f"      {name}: {email}")
  print("\n\n")


async def temp_big_test(links: list):
  for link in links:
    if link == "https://www.wellesley.edu/physics/people/faculty":
      pause = True
    await test(link)


# the following function grabs the input from a designated sheet.
# the input parameter is a path string to a valid excel sheet for input (see README for format).
# important: input is assumed to only have one sheet!
def read_input(input_filepath: str):
  output = {} # output is a dictionary from school names to lists of DeptLinks...
  if not path.exists(input_filepath):
    print("Could not read input excel spreadsheet: invalid path.")
    return output

  try:
    wb = load_workbook(filename=input_filepath, read_only=True)
  except Exception as e:
    print(f"Could not open input excel spreadsheet, raised: {str(e)}")
    return output

  # we assume that the workbook only has one sheet, so we can just grab the active one.
  sheet = wb.active
  # we now assume that the start row has been properly defined.
  row_num = INPUT_START_ROW
  while (not (sheet.cell(row=row_num, column=INPUT_UNIVERSITY_COL).value is None)):
    # we continue through the loop while there are still universities.
    university_name = sheet.cell(row=row_num, column=INPUT_UNIVERSITY_COL).value
    dept_links = []

    # we increment the row while there is still a department where there should be one.
    row_num += 1
    while (not (sheet.cell(row=row_num, column=INPUT_DEPARTMENT_COL).value is None)):
      dept = sheet.cell(row=row_num, column=INPUT_DEPARTMENT_COL).value
      link = sheet.cell(row=row_num, column=INPUT_LINK_COL).value
      dept_links.append(DeptLink(dept, link))
      row_num += 1
    
    # set the key in the dictionary to the correct value.
    output[university_name] = dept_links
    # we already hit a row with a blank department, which indicates we are ready for the next university name (if there is one, that is)
  
  wb.close()
  return output



# this function puts it all together: it takes in the input, attempts to scrape it, and writes to a sheet for the output.
async def run_session(input_filepath: str, output_filepath: str):
  schools_input = read_input(input_filepath)
  if len(schools_input) == 0:
    print("Could not start session, as no input was read in. Check formatting.")
    return
  
  if not isinstance(output_filepath, str):
    print("Could not run session, invalid output filepath: not a string.")
    return 

  # if the output_filepath is just a directory, we need to append a name.
  if path.isdir(output_filepath):
    output_filepath += f"/{DEFAULT_OUTPUT_NAME}({str(datetime.now())}).xlsx"
  else:
    # we need to see if it has the correct ending.
    if (len(output_filepath) < 5) or (output_filepath[-5:] != ".xlsx"):
      # file path is either too short or does not have the right ending.
      output_filepath += ".xlsx"

  # we can begin.
  print(f"Beginning session, writing to: {output_filepath}.")
  out = Workbook() 
  out.active.title = "Master Sheet"
  master = out["Master Sheet"]

  # first, we set all the title text on the master sheet.
  master.cell(row=1, column=1).value = "Successful Retrievals"
  master.cell(row=1, column=2).value = "Success count:"
  master.cell(row=1, column=3).value = 0

  master.cell(row=1, column=5).value = "Failed Retrievals"
  master.cell(row=1, column=6).value = "Fail count:"
  master.cell(row=1, column=7).value = 0

  master.cell(row=2, column=1).value = "University Name"
  master.cell(row=2, column=2).value = "Department"
  master.cell(row=2, column=3).value = "Number of Contacts"

  master.cell(row=2, column=5).value = "University Name"
  master.cell(row=2, column=6).value = "Department"
  master.cell(row=2, column=7).value = "Link"

  # do fills now.
  master.cell(row=1, column=1).fill = PatternFill(start_color=YELLOW_HEX, end_color=YELLOW_HEX, fill_type="solid")
  master.cell(row=1, column=3).fill = PatternFill(start_color=GREEN_HEX, end_color=GREEN_HEX, fill_type="solid")
  master.cell(row=1, column=4).fill = PatternFill(start_color=LIGHT_GRAY_HEX, end_color=LIGHT_GRAY_HEX, fill_type="solid")
  master.cell(row=2, column=4).fill = PatternFill(start_color=LIGHT_GRAY_HEX, end_color=LIGHT_GRAY_HEX, fill_type="solid")
  master.cell(row=1, column=5).fill = PatternFill(start_color=YELLOW_HEX, end_color=YELLOW_HEX, fill_type="solid")
  master.cell(row=1, column=7).fill = PatternFill(start_color=RED_HEX, end_color=RED_HEX, fill_type="solid")

  # now bolded text
  master.cell(row=1, column=3).font = Font(bold=True)
  master.cell(row=1, column=7).font = Font(bold=True)
  master.cell(row=2, column=1).font = Font(bold=True)
  master.cell(row=2, column=2).font = Font(bold=True)
  master.cell(row=2, column=3).font = Font(bold=True)
  master.cell(row=2, column=5).font = Font(bold=True)
  master.cell(row=2, column=6).font = Font(bold=True)
  master.cell(row=2, column=7).font = Font(bold=True)

  # now borders.
  for i in range(1, 8):
    if i == 4:
      continue # do not want to change this border...
    master.cell(row=2, column=i).border = Border(bottom=Side(style="thin"))

  master.cell(row=1, column=1).border = Border(left=Side(style="thin"))
  master.cell(row=2, column=1).border = Border(left=Side(style="thin"), bottom=Side(style="thin"))

  master.cell(row=1, column=3).border = Border(right=Side(style="thin"))
  master.cell(row=2, column=3).border = Border(right=Side(style="thin"), bottom=Side(style="thin"))

  master.cell(row=1, column=5).border = Border(left=Side(style="thin"))
  master.cell(row=2, column=5).border = Border(left=Side(style="thin"), bottom=Side(style="thin"))

  master.cell(row=1, column=7).border = Border(right=Side(style="thin"))
  master.cell(row=2, column=7).border = Border(right=Side(style="thin"), bottom=Side(style="thin"))
  try:
    out.save(output_filepath)
  except Exception as e:
    print(f"Could not save spreadsheet in setup, raised: {str(e)}")
    return 

  # now, we can finally begin to enter in the data as we get it.
  # we need to keep track of a bunch of stuff: lists of successful school names (for weird formatting stuff), the success row and success col we are on, individual rows for each page, and the number of successes and failures.
  succ_schools = []
  fail_schools = []
  succ_row_num = 3
  fail_row_num = 3
  succ_count = 0
  fail_count = 0

  for school_name, dept_links in schools_input.items():
    indiv_succ_row = 3 # row for success for an individual school.
    for dl in dept_links:
      dept = dl.department
      link = dl.link
      try:
        names_and_emails = await scrape_individual(link)
      except:
        names_and_emails = {} # indicates a failure.
      
      if len(names_and_emails) == 0:
        # we have failed.
        # first, check if it is not on the failure list.
        if not (school_name in fail_schools):
          fail_schools.append(school_name)
          master.cell(row=fail_row_num, column=5).value = school_name
          fail_row_num += 1 # so that we can enter the department one more time.
        
        fail_count += 1 # we failed.
        master.cell(row=fail_row_num, column=6).value = dept
        master.cell(row=fail_row_num, column=7).value = link
        fail_row_num += 1
        master.cell(row=1, column=7).value = fail_count
      else:
        # we have succeeded. 
        # first, check if this is not on the success list.
        if not (school_name in succ_schools):
          succ_schools.append(school_name)
          master.cell(row=succ_row_num, column=1).value = school_name 
          succ_row_num += 1
          
          # now, before we proceed, we need to make an individual page for this school.
          out.create_sheet(school_name)
          indiv = out[school_name]
          indiv.cell(row=1, column=1).value = school_name
          indiv.cell(row=2, column=1).value = "Department"
          indiv.cell(row=2, column=2).value = "Name"
          indiv.cell(row=2, column=3).value = "Email"

          # set colors and bolded text quickly.
          indiv.cell(row=1, column=1).fill = PatternFill(start_color=YELLOW_HEX, end_color=YELLOW_HEX, fill_type="solid")
          indiv.cell(row=2, column=1).font = Font(bold=True)
          indiv.cell(row=2, column=2).font = Font(bold=True)
          indiv.cell(row=2, column=3).font = Font(bold=True)

          # set borders.
          indiv.cell(row=1, column=1).border = Border(left=Side(style="thin"))
          indiv.cell(row=2, column=1).border = Border(left=Side(style="thin"), bottom=Side(style="thin"))
          indiv.cell(row=2, column=2).border = Border(bottom=Side(style="thin"))
          indiv.cell(row=2, column=3).border = Border(bottom=Side(style="thin"), right=Side(style="thin"))
          indiv.cell(row=1, column=3).border = Border(right=Side(style="thin"))

        # fill in the master sheet first.
        succ_count += 1
        master.cell(row=succ_row_num, column=2).value = dept
        master.cell(row=succ_row_num, column=3).value = len(names_and_emails)
        succ_row_num += 1
        master.cell(row=1, column=3).value = succ_count

        # we can now fill in all the info we need to in the individual page.
        indiv = out[school_name]
        indiv.cell(row=indiv_succ_row, column=1).value = dept
        indiv_succ_row += 1
        for name, email in names_and_emails.items():
          indiv.cell(row=indiv_succ_row, column=2).value = name
          indiv.cell(row=indiv_succ_row, column=3).value = email
          indiv_succ_row += 1
        
      # finally, we just update the grey in the master sheet for aesthetic purposes.
      for r in range(1, max(succ_row_num, fail_row_num) + 1):
        master.cell(row=r, column=4).fill = PatternFill(start_color=LIGHT_GRAY_HEX, end_color=LIGHT_GRAY_HEX, fill_type="solid")
      out.save(output_filepath)
      
  out.save(output_filepath)
  print("Done!")


if __name__ == "__main__":
  # we need to check for the correct number of arguments.
  if len(sys.argv) != 3:
    print("Usage: Python3 main.py input-excel-path output-excel-path.")
  else:
    asyncio.get_event_loop().run_until_complete(run_session(sys.argv[1], sys.argv[2]))
  
